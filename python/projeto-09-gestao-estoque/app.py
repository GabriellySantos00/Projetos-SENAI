from flask import Flask, render_template, request, redirect, url_for, send_file
import pymysql
import os
import openpyxl

app = Flask(__name__)

# Configuração da conexão com o banco de dados MySQL
def get_db_connection():
    try:
        conn = pymysql.connect(
            host='127.0.0.1',
            user='root',         
            password='dev@2025', 
            database='gestao_estoque',
            cursorclass=pymysql.cursors.DictCursor
        )
        return conn
    except Exception as err:
        print(f"Erro de conexão no Flask: {err}")
        return None

# Rota principal - Listagem de dados e conversão para renderização do gráfico
@app.route('/')
def index():
    conn = get_db_connection()
    if conn is None:
        return "Erro ao conectar ao banco de dados. Verifique o terminal.", 500
        
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM produtos")
    produtos = cursor.fetchall()
    cursor.close()
    conn.close()
    
    # Cast essencial para evitar que valores decimais quebrem a engine do JavaScript
    for produto in produtos:
        if produto.get('preco_unitario') is not None:
            produto['preco_unitario'] = float(produto['preco_unitario'])
    
    return render_template('index.html', produtos=produtos)

# Rota de inserção de novos ativos
@app.route('/criar', methods=['POST'])
def criar():
    nome = request.form['nome']
    categoria = request.form['categoria']
    quantidade = int(request.form['quantidade'])
    preco_unitario = float(request.form['preco_unitario'])
    
    conn = get_db_connection()
    if conn is not None:
        cursor = conn.cursor()
        sql = "INSERT INTO produtos (nome, categoria, quantidade, preco_unitario) VALUES (%s, %s, %s, %s)"
        cursor.execute(sql, (nome, categoria, quantidade, preco_unitario))
        conn.commit()
        cursor.close()
        conn.close()
        
    return redirect(url_for('index'))

# Rota de edição dinâmica em tempo real
@app.route('/editar/<int:id>', methods=['POST'])
def editar(id):
    nova_quantidade = int(request.form['quantidade'])
    novo_preco = float(request.form['preco_unitario'])
    
    conn = get_db_connection()
    if conn is not None:
        cursor = conn.cursor()
        sql = "UPDATE produtos SET quantidade = %s, preco_unitario = %s WHERE id = %s"
        cursor.execute(sql, (nova_quantidade, novo_preco, id))
        conn.commit()
        cursor.close()
        conn.close()
        
    return redirect(url_for('index'))

# Rota de remoção lógica de produtos
@app.route('/deletar/<int:id>')
def deletar(id):
    conn = get_db_connection()
    if conn is not None:
        cursor = conn.cursor()
        sql = "DELETE FROM produtos WHERE id = %s"
        cursor.execute(sql, (id,))
        conn.commit()
        cursor.close()
        conn.close()
        
    return redirect(url_for('index'))

# Rota para sincronização dinâmica e download do arquivo Excel (.xlsx)
@app.route('/planilha')
def gerar_planilha():
    nome_arquivo = 'pasta2.xlsx' if os.path.exists('pasta2.xlsx') else 'Pasta2.xlsx'
    
    conn = get_db_connection()
    if conn is None:
        return "Erro ao carregar dados para exportação.", 500
        
    cursor = conn.cursor()
    cursor.execute("SELECT nome, categoria, quantidade, preco_unitario FROM produtos")
    produtos = cursor.fetchall()
    cursor.close()
    conn.close()

    if os.path.exists(nome_arquivo):
        wb = openpyxl.load_workbook(nome_arquivo)
        ws = wb.active
        # Limpeza segura do histórico de linhas mantendo os cabeçalhos fixados na linha 1
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estoque"
        # Escreve os cabeçalhos padrão se o arquivo não existir fisicamente no diretório
        ws.append(["Nome", "Categoria", "Quantidade", "Preco Unitario"])

    # Força a exibição das linhas de grade para visualização corporativa padrão
    ws.views.sheetView[0].showGridLines = True

    # Escreve em tempo real todos os produtos cadastrados pelo usuário no banco
    for prod in produtos:
        ws.append([
            prod['nome'],
            prod['categoria'],
            prod['quantidade'],
            float(prod['preco_unitario'])
        ])

    wb.save(nome_arquivo)
    return send_file(nome_arquivo, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)